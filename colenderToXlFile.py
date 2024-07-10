# -*- coding: utf-8 -*-



import calendar
from openpyxl import Workbook
from openpyxl.styles import Alignment

# Создаем новую книгу Excel
wb = Workbook()
ws = wb.active

# Устанавливаем название листа
year = 2024  # Замените на нужный год

ws.title =  f"Календарь {year}"

# Генерируем календарь на текущий год
months = [f'Январь  {year}', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь', 'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь']
days = ['Пн', 'Вт', 'Ср', 'Чт', 'Пт', 'Сб', 'Вс']

# Добавляем месяцы и дни недели
for i, month in enumerate(months, start=1):
    ws.merge_cells(start_row=(i - 1) * 8 + 1, start_column=1, end_row=(i - 1) * 8 + 1, end_column=7)
    cell = ws.cell(row=(i - 1) * 8 + 1, column=1)
    cell.value = month
    cell.alignment = Alignment(horizontal='center')

    for j, day in enumerate(days, start=1):
        cell = ws.cell(row=(i - 1) * 8 + 2, column=j)
        cell.value = day
        cell.alignment = Alignment(horizontal='center')

    # Генерируем дни для каждого месяца
    month_days = calendar.monthcalendar(year, i)
    for week in month_days:
        row = month_days.index(week) + (i - 1) * 8 + 3
        for j, day in enumerate(week, start=1):
            if day != 0:
                cell = ws.cell(row=row, column=j)
                cell.value = day

# Сохраняем файл
wb.save(f'calendar{year}.xlsx')

